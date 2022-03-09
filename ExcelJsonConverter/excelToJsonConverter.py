import openpyxl
import json

book = openpyxl.load_workbook('DialogVorlage.xlsx')

def exportFrom(activeSheet):
    sheet = activeSheet    
    text = "["
    
    for i in range(1 , 500):
        keyObj = sheet.cell(row=i,column=1)
        keyVal = keyObj.value
    
        if keyVal == "background":
            valObj = sheet.cell(row=i,column=2)
            valueVal = valObj.value
            text += "{\"" + keyVal + "\":"
            text += "\"" + valueVal + "\"}"
        elif keyVal == "text":    
            valObj = sheet.cell(row=i,column=2)
            valueVal = valObj.value        
            tempKeyObj = sheet.cell(row=i-1,column=1)
            tempVal = tempKeyObj.value
            if tempVal == None:
                text += "{\"" + keyVal + "\":"
                text += "\"" + valueVal + "\"}"
        elif keyVal == "character":
            valObj = sheet.cell(row=i,column=2)
            valueVal = valObj.value
            text += "{\"" + keyVal + "\":"
            text += "\"" + valueVal + "\","
            
            tempKeyObj = sheet.cell(row=i+1,column=1)
            tempVal = tempKeyObj.value
            valObj = sheet.cell(row=i+1,column=2)
            valueVal = valObj.value
            text += "\"" + tempVal + "\":"
            text += "\"" + valueVal + "\","
            
            tempKeyObj = sheet.cell(row=i+2,column=1)
            tempVal = tempKeyObj.value
            valObj = sheet.cell(row=i+2,column=2)
            valueVal = valObj.value
            text += "\"" + tempVal + "\":"
            text += "\"" + valueVal + "\","
            
            tempKeyObj = sheet.cell(row=i+3,column=1)
            tempVal = tempKeyObj.value
            valObj = sheet.cell(row=i+3,column=2)
            valueVal = valObj.value
            text += "\"" + tempVal + "\":"
            text += "\"" + valueVal + "\"}"
        
        elif keyVal == "question":
            valObj = sheet.cell(row=i,column=2)
            valueVal = valObj.value
            
            text += "{\"" + keyVal + "\":"
            text += "\"" + valueVal + "\","
            text += "\"options\":["
            tempKeyObj = sheet.cell(row=i+1,column=1)
            tempVal = tempKeyObj.value
            valObj = sheet.cell(row=i+1,column=2)
            valueVal = valObj.value
            text += "{\"label\":\"" + tempVal + "\",\"value\":\"" + str(valueVal) + "\"},"
            tempKeyObj = sheet.cell(row=i+2,column=1)
            tempVal = tempKeyObj.value
            valObj = sheet.cell(row=i+2,column=2)
            valueVal = valObj.value
            text += "{\"label\":\"" + tempVal + "\",\"value\":\"" + str(valueVal) + "\"}],"
            tempKeyObj = sheet.cell(row=i+3,column=1)
            tempVal = tempKeyObj.value
            valObj = sheet.cell(row=i+3,column=2)
            valueVal = valObj.value
            text += "\"" + tempVal + "\":"
            text += "\"" + str(valueVal) + "\"}"

        elif keyVal == "action":
            valObj = sheet.cell(row=i, column=2)
            valueVal = valObj.value
            text += "{\"" + keyVal + "\":"
            text += "\"" + valueVal + "\"}"

        elif keyVal == "scene":
            valObj = sheet.cell(row=i, column=2)
            valueVal = valObj.value
            text += "{\"" + keyVal + "\":"
            text += "\"" + valueVal + "\"}"
        
        elif keyVal == None:
            tBO = sheet.cell(row=i-1,column=1)
            tBV = tBO.value
            
            tAO = sheet.cell(row=i+1,column=1)
            tAV = tAO.value
            
            if tAV != None:
                text += ","
    
    text += "]"


    with open(sheet.title + ".json", 'w', encoding='utf-8') as f:
        json.dump(text.replace("\\", ""), f, ensure_ascii=False, indent=4)
    
    print(text)

for sh in book.worksheets:
    exportFrom(sh)
